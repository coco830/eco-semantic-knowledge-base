import csv
import json
import re
from pathlib import Path

from openpyxl import load_workbook
from pypdf import PdfReader


ROOT = Path(__file__).resolve().parent
GB_XLSX = ROOT / "2017国民经济行业分类注释.xlsx"
PERMIT_PDF = ROOT / "固定污染源排污许可分类管理名录(2019年版).pdf"


CURRENT_FIRST_SECTIONS = {f"S{i:02d}" for i in range(1, 18)} | {"RADIATION"}
CURRENT_MONTHLY_SECTIONS = {f"M{i:02d}" for i in range(1, 10)} | {"RADIATION"}
MAPPED_BUT_NOT_CURRENT_MONTHLY = {f"M{i:02d}" for i in range(10, 18)}


OPEN_QUESTIONS = [
    {
        "question_id": "OQ-001",
        "topic": "GB4754_CODE_NAME_CONFLICT",
        "object_type": "industry_rule",
        "object_id": "GB4754_1512_*",
        "issue": "现有样板将1512写为啤酒制造，但本地GB/T 4754-2017注释中1512为白酒制造，啤酒制造为1513。",
        "impact": "不能进入正式规则；会导致行业召回和许可条件错误。",
        "source_basis": "2017国民经济行业分类注释.xlsx 行1777/1787",
        "owner_role": "ETO",
        "status": "OPEN",
        "recommended_action": "确认是否将啤酒样板代码改为1513，或将现有1512改成白酒制造样板。",
    },
    {
        "question_id": "OQ-002",
        "topic": "REPRESENTATIVE_SUBCLASS_SCOPE",
        "object_type": "industry_rule",
        "object_id": "GB4754_1521_*",
        "issue": "1521为碳酸饮料制造，样板名称写饮料制造；若业务意图为152饮料制造，不应由单一小类承载。",
        "impact": "RAG召回可能把其他饮料小类漏掉或误归到碳酸饮料。",
        "source_basis": "2017国民经济行业分类注释.xlsx；固定污染源排污许可分类管理名录(2019年版)第22条",
        "owner_role": "ESO/ETO",
        "status": "OPEN",
        "recommended_action": "按152中类展开到各四位小类，或明确1521仅作为碳酸饮料样板。",
    },
    {
        "question_id": "OQ-003",
        "topic": "REPRESENTATIVE_SUBCLASS_SCOPE",
        "object_type": "industry_rule",
        "object_id": "GB4754_2211_KEY",
        "issue": "2211为木竹浆制造，样板名称写纸浆制造；2019名录第36条按221纸浆制造中类管理。",
        "impact": "非木竹浆纸浆小类覆盖范围不清。",
        "source_basis": "2017国民经济行业分类注释.xlsx；固定污染源排污许可分类管理名录(2019年版)第36条",
        "owner_role": "ETO",
        "status": "OPEN",
        "recommended_action": "确认用221中类规则下挂所有纸浆小类，还是仅保留2211样板。",
    },
    {
        "question_id": "OQ-004",
        "topic": "REPRESENTATIVE_SUBCLASS_SCOPE",
        "object_type": "industry_rule",
        "object_id": "GB4754_2231_REGISTRATION",
        "issue": "2019名录第38条按223纸制品制造中类判断，2231只是代表性小类，且许可类型取决于是否有工业废水或废气。",
        "impact": "登记/简化边界不能只靠2231默认确定。",
        "source_basis": "2017国民经济行业分类注释.xlsx；固定污染源排污许可分类管理名录(2019年版)第38条",
        "owner_role": "ESO/ETO",
        "status": "OPEN",
        "recommended_action": "补充是否有工业废水/废气确认问题，并按223中类展开候选。",
    },
    {
        "question_id": "OQ-005",
        "topic": "GENERAL_PROCESS_TRIGGER",
        "object_type": "industry_rule",
        "object_id": "GB4754_1371_*",
        "issue": "蔬菜加工样板许可类型依赖通用工序，不能仅靠行业代码定KEY/SIMPLIFIED/REGISTRATION。",
        "impact": "会把通用工序触发条件误写成行业默认画像。",
        "source_basis": "固定污染源排污许可分类管理名录(2019年版)第15条",
        "owner_role": "ESO/ETO",
        "status": "OPEN",
        "recommended_action": "增加锅炉、工业炉窑、表面处理、水处理等通用工序确认问题。",
    },
    {
        "question_id": "OQ-006",
        "topic": "GENERAL_PROCESS_TRIGGER",
        "object_type": "industry_rule",
        "object_id": "GB4754_3311_*",
        "issue": "金属结构制造样板许可类型依赖通用工序和喷涂/酸洗/表面处理事实。",
        "impact": "会把喷涂、酸洗、抛光、热浸镀等现场事实漏为行业固有属性。",
        "source_basis": "固定污染源排污许可分类管理名录(2019年版)第80条",
        "owner_role": "ESO/ETO",
        "status": "OPEN",
        "recommended_action": "按金属加工场景建立确认问题，不直接默认许可类型。",
    },
    {
        "question_id": "OQ-007",
        "topic": "TEMPLATE_SECTION_CANDIDATES",
        "object_type": "inspection_recommendation",
        "object_id": "S18/S19/S20/S21/S22/S23/S24/M18/M19",
        "issue": "排查建议中存在当前模板没有的建议章节，且部分语义曾在模板瘦身中移除。",
        "impact": "不得直接生成检查模板或接入小程序运行时。",
        "source_basis": "用户确认：知识库v0.1样板，不直接接入运行时。",
        "owner_role": "Product/ETO",
        "status": "OPEN",
        "recommended_action": "将这些章节标记为KNOWLEDGE_CANDIDATE_SECTION，只用于知识库建议和未来评审。",
    },
]


def write_csv(path, headers, rows):
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=headers)
        writer.writeheader()
        writer.writerows(rows)


def build_industry_catalog():
    wb = load_workbook(GB_XLSX, read_only=True, data_only=True)
    ws = wb["Sheet1"]
    rows = []
    seen = set()
    current_section = ""
    current_section_code = ""
    current_division_code = ""
    current_division_name = ""
    current_group_code = ""
    current_group_name = ""

    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        cells = [str(v).strip() if v is not None else "" for v in row[:5]]
        if not any(cells):
            continue
        col_a = cells[0]
        col_b = cells[1]
        name = cells[3] or cells[4]

        if re.fullmatch(r"[A-Z]", col_a):
            current_section_code = col_a
            current_section = name
            continue
        if re.fullmatch(r"\d{2}", col_a):
            current_division_code = col_a
            current_division_name = name
            current_group_code = ""
            current_group_name = ""
            item = {
                "level": "division",
                "section_code": current_section_code,
                "section_name": current_section,
                "division_code": col_a,
                "division_name": name,
                "group_code": "",
                "group_name": "",
                "class_code": "",
                "class_name": "",
                "gb_source_row": row_idx,
                "status": "BASE_CATALOG",
                "notes": "",
            }
            key = (item["level"], item["division_code"], item["group_code"], item["class_code"])
            if key not in seen:
                seen.add(key)
                rows.append(item)
            continue
        if re.fullmatch(r"\d{3}", col_a):
            current_group_code = col_a
            current_group_name = name
            item = {
                "level": "group",
                "section_code": current_section_code,
                "section_name": current_section,
                "division_code": current_division_code,
                "division_name": current_division_name,
                "group_code": col_a,
                "group_name": name,
                "class_code": "",
                "class_name": "",
                "gb_source_row": row_idx,
                "status": "BASE_CATALOG",
                "notes": "",
            }
            key = (item["level"], item["division_code"], item["group_code"], item["class_code"])
            if key not in seen:
                seen.add(key)
                rows.append(item)
        class_code = ""
        if re.fullmatch(r"\d{4}", col_b):
            class_code = col_b
        elif re.fullmatch(r"\d{4}", col_a):
            class_code = col_a

        if class_code:
            class_name = name
            if not current_group_code and len(class_code) >= 3:
                current_group_code = class_code[:3]
            item = {
                "level": "class",
                "section_code": current_section_code,
                "section_name": current_section,
                "division_code": current_division_code,
                "division_name": current_division_name,
                "group_code": current_group_code,
                "group_name": current_group_name,
                "class_code": class_code,
                "class_name": class_name,
                "gb_source_row": row_idx,
                "status": "BASE_CATALOG",
                "notes": "same_row_group_and_class" if re.fullmatch(r"\d{3}", col_a) and re.fullmatch(r"\d{4}", col_b) else "",
            }
            key = (item["level"], item["division_code"], item["group_code"], item["class_code"])
            if key not in seen:
                seen.add(key)
                rows.append(item)
    return rows


def build_permit_catalog_draft():
    reader = PdfReader(PERMIT_PDF)
    rows = []
    active_major = ""
    for page_no, page in enumerate(reader.pages, start=1):
        text = page.extract_text() or ""
        if page_no < 4:
            continue
        compact = " ".join(text.split())
        for m in re.finditer(r"([一二三四五六七八九十]+、[^0-9]+?\s\d{2})", compact):
            active_major = m.group(1)
        # Entry parsing is intentionally conservative: raw text is retained for audit.
        for m in re.finditer(r"(?<!\d)(\d{1,3})\s+(.{8,260}?)(?=\s+\d{1,3}\s+|$)", compact):
            entry_no = int(m.group(1))
            if not (1 <= entry_no <= 112):
                continue
            raw = m.group(2).strip()
            if "序号 行业类别" in raw or len(raw) < 8:
                continue
            codes = sorted(set(re.findall(r"\b\d{3,4}\b", raw)))
            rows.append({
                "catalog_entry_no": entry_no,
                "major_category_text": active_major,
                "industry_category_text": raw,
                "gb_code_fragments": ";".join(codes),
                "key_management_condition": "SEE_RAW_TEXT",
                "simplified_management_condition": "SEE_RAW_TEXT",
                "registration_management_condition": "SEE_RAW_TEXT",
                "source_page": page_no,
                "extraction_confidence": "LOW",
                "audit_status": "RAW_TEXT_NEEDS_MANUAL_SPLIT",
                "notes": "PDF table extraction retains raw row text first; split key/simplified/registration conditions before formal rule generation.",
            })
    dedup = {}
    for row in rows:
        dedup.setdefault(row["catalog_entry_no"], row)
    return [dedup[k] for k in sorted(dedup)]


def section_status(code, inspection_type):
    parts = code.split("/")
    statuses = []
    for part in parts:
        part = part.strip()
        if inspection_type == "FIRST" and part in CURRENT_FIRST_SECTIONS:
            statuses.append("CURRENT_TEMPLATE_SECTION")
        elif inspection_type == "MONTHLY" and part in CURRENT_MONTHLY_SECTIONS:
            statuses.append("CURRENT_TEMPLATE_SECTION")
        elif inspection_type == "MONTHLY" and part in MAPPED_BUT_NOT_CURRENT_MONTHLY:
            statuses.append("MAPPED_NOT_CURRENT_TEMPLATE_SECTION")
        else:
            statuses.append("KNOWLEDGE_CANDIDATE_SECTION")
    if all(s == "CURRENT_TEMPLATE_SECTION" for s in statuses):
        return "CURRENT_TEMPLATE_SECTION"
    if any(s == "KNOWLEDGE_CANDIDATE_SECTION" for s in statuses):
        return "KNOWLEDGE_CANDIDATE_SECTION"
    return "MAPPED_NOT_CURRENT_TEMPLATE_SECTION"


def annotate_inspection_recommendations():
    path = ROOT / "inspection_item_recommendations.csv"
    if not path.exists():
        return []
    with path.open(encoding="utf-8-sig", newline="") as f:
        rows = list(csv.DictReader(f))
    for row in rows:
        status = section_status(row["suggested_section_code"], row["inspection_type"])
        row["section_status"] = status
        row["runtime_status"] = "NOT_FOR_RUNTIME_V0_1" if status != "CURRENT_TEMPLATE_SECTION" else "REVIEW_BEFORE_RUNTIME"
        row["knowledge_base_role"] = "candidate_recommendation"
    headers = list(rows[0].keys()) if rows else []
    write_csv(ROOT / "inspection_item_recommendations.csv", headers, rows)
    candidate_rows = []
    for row in rows:
        candidate_rows.append({
            "scenario_id": row["scenario_id"],
            "inspection_type": row["inspection_type"],
            "candidate_section_codes": row["suggested_section_code"],
            "candidate_subchapter_name": row["suggested_subchapter"],
            "item_title": row["item_title"],
            "check_points": row["check_points"],
            "evidence_policy": row["evidence_policy"],
            "photo_points": row["photo_points"],
            "answer_policy": row["answer_policy"],
            "candidate_severity": row["default_severity"],
            "candidate_deduct": row["default_deduct"],
            "activation_condition": row["show_if_condition"],
            "section_status": row["section_status"],
            "runtime_status": "CANDIDATE_ONLY",
            "candidate_reason": "v0.1 knowledge-base recommendation; review current template sections before runtime use.",
        })
    if candidate_rows:
        write_csv(
            ROOT / "inspection_candidate_recommendations_v0_1.csv",
            list(candidate_rows[0].keys()),
            candidate_rows,
        )
    return rows


def write_open_questions():
    headers = list(OPEN_QUESTIONS[0].keys())
    write_csv(ROOT / "open_questions.csv", headers, OPEN_QUESTIONS)
    lines = [
        "# Open Questions",
        "",
        "这些问题必须在规则进入正式运行时或批量扩展前处理。`OPEN` 不阻塞 v0.1 样板研究，但阻塞正式接入。",
        "",
    ]
    for q in OPEN_QUESTIONS:
        lines.extend([
            f"## {q['question_id']} {q['topic']}",
            "",
            f"- 对象：`{q['object_id']}`",
            f"- 问题：{q['issue']}",
            f"- 影响：{q['impact']}",
            f"- 来源：{q['source_basis']}",
            f"- 建议动作：{q['recommended_action']}",
            "",
        ])
    (ROOT / "open_questions.md").write_text("\n".join(lines), encoding="utf-8")


def write_manifest_and_plan(industry_rows, permit_rows, inspection_rows):
    table_validation_path = ROOT / "permit_management_catalog_table_cells_validation.json"
    table_validation = None
    if table_validation_path.exists():
        table_validation = json.loads(table_validation_path.read_text(encoding="utf-8"))
    manifest = {
        "knowledge_base_version": "v0.1-sample",
        "positioning": "auditable knowledge-base sample for scenario ontology, score13 semantic mapping, evidence and photo guidance; not runtime input.",
        "runtime_integration": "disabled",
        "generated_outputs": {
            "industry_catalog_base.csv": len(industry_rows),
            "permit_management_catalog_draft.csv": len(permit_rows),
            "permit_management_catalog_table_cells.csv": table_validation["entry_count"] if table_validation else 0,
            "open_questions.csv": len(OPEN_QUESTIONS),
            "inspection_item_recommendations.csv": len(inspection_rows),
            "inspection_candidate_recommendations_v0_1.csv": len(inspection_rows),
        },
        "quality_gates_before_runtime": [
            "resolve all OPEN questions that affect industry code or permit type",
            "use permit_management_catalog_table_cells.csv as the raw audit source and verify catalog entries 1-112",
            "split permit raw rows into key/simplified/registration conditions with page references",
            "map candidate sections to current template sections or explicitly approve new sections",
            "run duplicate scenario and source_basis audit",
            "ESO/ETO confirms final enterprise profile from EIA, approval, permit, ledger and site facts",
        ],
        "permit_catalog_table_cell_validation": table_validation or "NOT_GENERATED",
    }
    (ROOT / "knowledge_base_v0_1_manifest.json").write_text(
        json.dumps(manifest, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )

    (ROOT / "runtime_integration_boundary.md").write_text(
        "# Runtime Integration Boundary\n\n"
        "本批产物定位为 `knowledge_base_version = v0.1-sample`，不直接接入小程序或 EcoCheck 运行时。\n\n"
        "## 可以使用\n\n"
        "- 作为图谱/RAG 的样板语料和字段设计参考。\n"
        "- 作为 ESO/ETO 人工确认问题和现场拍照提示的候选来源。\n"
        "- 作为后续行业批量扩展、去重和质量抽检的底座。\n\n"
        "## 不可以使用\n\n"
        "- 不得把 `industry_scenario_rules.json` 直接写入企业最终画像。\n"
        "- 不得把 `inspection_item_recommendations.csv` 中的 `KNOWLEDGE_CANDIDATE_SECTION` 直接生成检查模板。\n"
        "- 优先把 `inspection_candidate_recommendations_v0_1.csv` 当作 v0.1 候选检查项视图；其中 `candidate_*`、`activation_condition`、`runtime_status=CANDIDATE_ONLY` 都是防误接语义。\n"
        "- 不得因行业代码和许可类型命中，就跳过环评、批复、排污许可、台账和现场事实确认。\n\n"
        "## 候选章节口径\n\n"
        "`S18/S19/S20/S21/S22/S23/S24/M18/M19` 等章节为知识库候选章节。"
        "它们表达的是未来可能增强的场景子章，不代表当前 EcoCheck 模板已经存在这些章节。\n",
        encoding="utf-8",
    )

    (ROOT / "README_KNOWLEDGE_BASE_V0_1.md").write_text(
        "# 行业产污场景知识库 v0.1 样板\n\n"
        "## 定位\n\n"
        "本目录当前产物是知识库 v0.1 样板，不是运行时规则库。它用于 RAG、环保语义图谱、规则治理、首轮企业画像预填、现场确认问题和拍照提示设计。\n\n"
        "行业代码和排污许可类型只作为默认召回入口；最终企业画像必须由 ESO/ETO 根据环评、批复、排污许可证、台账和现场事实确认。\n\n"
        "## 核心文件\n\n"
        "- `scenario_templates.json`：可复用产污场景模板。\n"
        "- `industry_scenario_rules.json`：31 条样板行业规则到场景模板的候选召回。\n"
        "- `industry_catalog_base.csv`：GB/T 4754-2017 全量层级底座。\n"
        "- `permit_management_catalog_draft.csv`：2019 排污许可名录原文草表，需人工拆分三类管理条件。\n"
        "- `permit_management_catalog_table_cells.csv`：2019 排污许可名录表格级审计中间表，已验证 1-112 连续时可作为 raw audit source。\n"
        "- `permit_management_catalog_table_cells_validation.json`：表格级抽取验证结果。\n"
        "- `scenario_to_score13_mapping.csv`：场景到 EcoCheck S01-S13 的语义映射。\n"
        "- `inspection_candidate_recommendations_v0_1.csv`：候选检查项、证据和拍照点，不直接接模板。\n"
        "- `open_questions.csv` / `open_questions.md`：阻塞运行时或需治理确认的问题。\n"
        "- `runtime_integration_boundary.md`：运行时禁用边界。\n\n"
        "## 当前硬边界\n\n"
        "- 不自动扣分。\n"
        "- 不自动创建正式检查项或报告章节。\n"
        "- 不把 `UNKNOWN` 当成 `YES` 或 `NO`。\n"
        "- 不直接改 EcoCheck S01-S13 报告口径。\n"
        "- 不把候选章节 `S18/S19/S20/S21/S22/S23/S24/M18/M19` 当作现行模板章节。\n\n"
        "## 已知限制\n\n"
        "`permit_management_catalog_draft.csv` 由 PDF 文本保守抽取生成，只用于定位原文和字段设计草稿。"
        "当前草表未达到 1-112 条连续完整抽取，不可用于正式规则。"
        "`permit_management_catalog_table_cells.csv` 是更可靠的表格级 raw audit source，但仍未做阈值、星号、通用工序、重点排污单位等条件规则化，因此也不能直接生成 `permit_type`。\n\n"
        "## 下一步\n\n"
        "按 `expansion_batch_plan.md` 先跑 Batch 1：农副食品、食品、酒饮料茶、造纸纸制品、印刷相邻行业。每批先复用现有 `SCN_*` 模板，无法确认时标 `NEED_CONFIRM`。\n",
        encoding="utf-8",
    )

    (ROOT / "expansion_batch_plan.md").write_text(
        "# 剩余行业底座扩展计划\n\n"
        "## 原则\n\n"
        "- 先复用现有场景模板，不为每个行业单独造规则。\n"
        "- 行业代码和排污许可类型只做默认召回入口，最终企业画像由 ESO/ETO 根据环评、批复、排污许可、台账和现场事实确认。\n"
        "- 对通用工序触发、代表性小类、许可条件无法拆分的内容标 `NEED_CONFIRM`。\n"
        "- 每批扩展必须抽检：行业代码层级、名录条目、场景复用、open questions、重复证据点。\n\n"
        "## Batch 1：已有样板相邻高价值行业\n\n"
        "扩展 15 酒饮料、13 农副食品、14 食品、22 造纸纸制品、23 印刷。目标是修正 `1512/1513`，把中类规则展开为四位小类候选，并复用废水、VOCs、危废、在线监测、一般固废模板。\n\n"
        "## Batch 2：涉 VOCs/表面处理/喷涂行业\n\n"
        "扩展 17-21 纺织、服装、皮革、木材、家具，以及 33-38 装备制造相关条目。重点只做默认候选：前处理/染色/喷涂/胶粘剂/磷化/酸洗等全部进入 `NEED_CONFIRM` 触发问题。\n\n"
        "## Batch 3：高水量、高粉尘、高风险单元行业\n\n"
        "扩展 03 畜牧、13 屠宰/水产/淀粉、25-31 石化/化工/医药/橡塑/非金属矿物、30 水泥相邻条目。优先沉淀废水、粉尘、罐区、LDAR、土壤地下水、应急事故水模板复用。\n\n"
        "## 质量抽检规则\n\n"
        "- 每批至少抽检 10% 行业小类，且覆盖 HIGH/MEDIUM/LOW/NEED_CONFIRM。\n"
        "- 每条规则必须有 `source_basis`、`confidence`、`confirmation_questions`。\n"
        "- 同一证据点只能在多个章节重复出现时必须说明目的差异。\n"
        "- 候选章节和当前模板章节必须分开标识，禁止候选章节直接进入运行时。\n",
        encoding="utf-8",
    )


def main():
    industry_rows = build_industry_catalog()
    write_csv(
        ROOT / "industry_catalog_base.csv",
        ["level", "section_code", "section_name", "division_code", "division_name", "group_code", "group_name", "class_code", "class_name", "gb_source_row", "status", "notes"],
        industry_rows,
    )

    permit_rows = build_permit_catalog_draft()
    write_csv(
        ROOT / "permit_management_catalog_draft.csv",
        ["catalog_entry_no", "major_category_text", "industry_category_text", "gb_code_fragments", "key_management_condition", "simplified_management_condition", "registration_management_condition", "source_page", "extraction_confidence", "audit_status", "notes"],
        permit_rows,
    )

    inspection_rows = annotate_inspection_recommendations()
    write_open_questions()
    write_manifest_and_plan(industry_rows, permit_rows, inspection_rows)

    print(json.dumps({
        "industry_catalog_rows": len(industry_rows),
        "industry_class_rows": sum(1 for r in industry_rows if r["level"] == "class"),
        "permit_catalog_draft_rows": len(permit_rows),
        "open_questions": len(OPEN_QUESTIONS),
        "inspection_recommendations_annotated": len(inspection_rows),
        "candidate_recommendations_written": bool(inspection_rows),
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
