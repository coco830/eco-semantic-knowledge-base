import csv
import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Protection
from openpyxl.worksheet.datavalidation import DataValidation


ROOT = Path(__file__).resolve().parent
FINAL_STATE = "NOT_FOR_RUNTIME_CANDIDATE_KB_ONLY"
RUNTIME_INTEGRATION = "disabled"

LABELS = [
    "CONFIRM_APPLIES",
    "CONFIRM_MAY_APPLY",
    "CONFIRM_NOT_APPLY",
    "NEED_EIA_CONFIRM",
    "NEED_PERMIT_CONFIRM",
    "NEED_SITE_CONFIRM",
    "REJECT_CANDIDATE",
    "MERGE_DUPLICATE",
    "NEED_RULE_FIX",
]

DECISION_CONFIDENCE = ["HIGH", "MEDIUM", "LOW", "NEED_CONFIRM"]
REVIEWER_ROLES = ["ESO", "ETO", "DomainOwner", "Product", "TechLead"]
EDITABLE_FIELDS = [
    "human_review_label",
    "human_reviewer",
    "human_reviewer_role",
    "reviewed_at",
    "human_review_notes",
    "review_basis",
    "evidence_refs",
    "decision_confidence",
]


def read_csv(path):
    with path.open(encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))


def read_json(path):
    return json.loads(path.read_text(encoding="utf-8"))


def read_jsonl(path):
    rows = []
    with path.open(encoding="utf-8") as f:
        for line in f:
            if line.strip():
                rows.append(json.loads(line))
    return rows


def write_csv(path, rows, fields=None):
    tmp = path.with_name(f".{path.name}.tmp")
    if fields is None:
        fields = list(rows[0].keys()) if rows else []
    with tmp.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fields)
        writer.writeheader()
        writer.writerows(rows)
    tmp.replace(path)


def write_json(path, data):
    tmp = path.with_name(f".{path.name}.tmp")
    tmp.write_text(json.dumps(data, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    tmp.replace(path)


def parse_json(value, default=None):
    if default is None:
        default = []
    if isinstance(value, (list, dict)):
        return value
    if not value:
        return default
    try:
        return json.loads(value)
    except json.JSONDecodeError:
        return default


def load_graph_refs():
    refs = {}
    for node in read_jsonl(ROOT / "graph_nodes_v0_5.jsonl"):
        if node.get("node_type") != "ApplicabilityRelation":
            continue
        rel_id = node["natural_key"]
        refs[rel_id] = {
            "open_question_refs": node.get("open_question_refs", []),
            "risk_refs": node.get("risk_refs", []),
        }
    return refs


def classify_review(row, refs):
    gate = row["gate_status"]
    source = row["relation_source"]
    entry = row["entry_no"]
    risk_refs = refs.get("risk_refs", [])
    open_refs = refs.get("open_question_refs", [])
    if "GENERAL_PROCESS_TRIGGER" in source or entry in {"109", "110", "111", "112"}:
        return "P0", "GENERAL_PROCESS_109_112_REVIEW"
    if "RISK-V041-005" in risk_refs or "V04_RUNTIME_APPROVAL_GATE_001" in open_refs:
        return "P0", "RUNTIME_BOUNDARY_REVIEW"
    if source == "DIVISION_CONTEXT":
        return "P1", "DIVISION_CONTEXT_RECALL_REVIEW"
    if gate == "NEED_EIA_OR_PERMIT_CONFIRM":
        return "P1", "EIA_PERMIT_CONFIRMATION"
    if gate == "MAY_APPLY":
        return "P1", "MAY_APPLY_CONFIRMATION"
    if gate == "APPLIES":
        return "P1", "APPLIES_EVIDENCE_REVIEW"
    if gate == "NOT_APPLY":
        return "P2", "NOT_APPLY_EXCLUSION_REVIEW"
    return "P2", "GENERAL_CANDIDATE_REVIEW"


def build_queue(context_rows, graph_refs):
    rows = []
    for row in context_rows:
        refs = graph_refs.get(row["candidate_relation_id"], {"open_question_refs": [], "risk_refs": []})
        priority, bucket = classify_review(row, refs)
        rows.append({
            "review_item_id": f"HRV07::{row['candidate_relation_id']}",
            "source_table": "all_context_applicability_review_v0_4_1.csv",
            "source_row_id": row["candidate_relation_id"],
            "candidate_relation_id": row["candidate_relation_id"],
            "industry_code": row["industry_code"],
            "industry_name": row["industry_name"],
            "entry_no": row["entry_no"],
            "target_management_condition": row["target_management_condition"],
            "raw_condition": row["raw_condition"],
            "normalized_condition": row["normalized_condition"],
            "gate_status": row["gate_status"],
            "gate_reason": row["gate_reason"],
            "relation_source": row["relation_source"],
            "source_basis": row["source_basis"],
            "confidence": row["confidence"],
            "blocking_flags": row["blocking_flags"],
            "confirmation_questions": row["confirmation_questions"],
            "related_scenario_ids": row["related_scenario_ids"],
            "open_question_refs": json.dumps(refs["open_question_refs"], ensure_ascii=False),
            "risk_refs": json.dumps(refs["risk_refs"], ensure_ascii=False),
            "review_priority": priority,
            "review_bucket": bucket,
            "final_state": FINAL_STATE,
            "runtime_integration": RUNTIME_INTEGRATION,
        })
    return rows


def build_worksheet(queue_rows):
    fields = list(queue_rows[0].keys()) + EDITABLE_FIELDS
    rows = []
    for row in queue_rows:
        out = dict(row)
        for field in EDITABLE_FIELDS:
            out[field] = ""
        rows.append(out)
    write_csv(ROOT / "human_review_worksheet_v0_7.csv", rows, fields)

    wb = Workbook()
    ws = wb.active
    ws.title = "Review Queue"
    ws.append(fields)
    for row in rows:
        ws.append([row.get(field, "") for field in fields])

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    editable_fill = PatternFill("solid", fgColor="FFF2CC")
    locked_fill = PatternFill("solid", fgColor="E7EEF7")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    editable_indexes = {fields.index(field) + 1 for field in EDITABLE_FIELDS}
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            if cell.column in editable_indexes:
                cell.fill = editable_fill
                cell.protection = Protection(locked=False)
            else:
                cell.fill = locked_fill
                cell.protection = Protection(locked=True)
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    widths = {
        "A": 34, "B": 34, "C": 28, "D": 28, "E": 12, "F": 24, "G": 10, "H": 16,
        "I": 42, "J": 46, "K": 16, "L": 28, "M": 24, "N": 40, "O": 12,
        "P": 35, "Q": 35, "R": 35, "S": 30, "T": 20, "U": 16, "V": 14,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    for idx in range(23, len(fields) + 1):
        ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = 24

    label_col = fields.index("human_review_label") + 1
    role_col = fields.index("human_reviewer_role") + 1
    conf_col = fields.index("decision_confidence") + 1
    label_dv = DataValidation(type="list", formula1=f'"{",".join(LABELS)}"', allow_blank=True)
    role_dv = DataValidation(type="list", formula1=f'"{",".join(REVIEWER_ROLES)}"', allow_blank=True)
    conf_dv = DataValidation(type="list", formula1=f'"{",".join(DECISION_CONFIDENCE)}"', allow_blank=True)
    ws.add_data_validation(label_dv)
    ws.add_data_validation(role_dv)
    ws.add_data_validation(conf_dv)
    label_dv.add(f"{ws.cell(row=2, column=label_col).coordinate}:{ws.cell(row=ws.max_row, column=label_col).coordinate}")
    role_dv.add(f"{ws.cell(row=2, column=role_col).coordinate}:{ws.cell(row=ws.max_row, column=role_col).coordinate}")
    conf_dv.add(f"{ws.cell(row=2, column=conf_col).coordinate}:{ws.cell(row=ws.max_row, column=conf_col).coordinate}")

    info = wb.create_sheet("Guidance")
    info.append(["Field", "Rule"])
    for field in EDITABLE_FIELDS:
        info.append([field, "Only these fields may be filled by reviewers; source fields are locked."])
    info.append(["Boundary", FINAL_STATE])
    info.append(["Runtime", "disabled; manual review is not runtime approval."])
    for cell in info[1]:
        cell.fill = header_fill
        cell.font = header_font
    info.column_dimensions["A"].width = 28
    info.column_dimensions["B"].width = 100
    ws.protection.sheet = True
    ws.protection.enable()
    wb.save(ROOT / "human_review_worksheet_v0_7.xlsx")
    return rows


def build_decision_schema():
    labels = []
    for label in LABELS:
        labels.append({
            "label": label,
            "requires_human_reviewer": True,
            "requires_human_review_notes": label.startswith("CONFIRM") or label in {"REJECT_CANDIDATE", "NEED_RULE_FIX"},
            "requires_review_basis": True,
            "requires_evidence_refs": label.startswith("CONFIRM") or label in {"REJECT_CANDIDATE", "NEED_RULE_FIX"},
            "runtime_effect": "NO_RUNTIME_EFFECT_IN_V0_7",
        })
    schema = {
        "schema_version": "v0.7",
        "final_state": FINAL_STATE,
        "runtime_integration": RUNTIME_INTEGRATION,
        "editable_fields": EDITABLE_FIELDS,
        "reviewer_roles": REVIEWER_ROLES,
        "decision_confidence_values": DECISION_CONFIDENCE,
        "labels": labels,
        "formalization_candidate_labels": ["CONFIRM_APPLIES", "CONFIRM_NOT_APPLY", "CONFIRM_MAY_APPLY"],
        "blocks_runtime_labels": ["NEED_EIA_CONFIRM", "NEED_PERMIT_CONFIRM", "NEED_SITE_CONFIRM", "REJECT_CANDIDATE", "MERGE_DUPLICATE", "NEED_RULE_FIX"],
        "second_approval_required_for": ["runtime_integration", "formal_permit_type", "formal_inspection_template", "auto_deduct", "score13_report_dimension_change"],
    }
    write_json(ROOT / "human_review_decision_schema_v0_7.json", schema)
    lines = [
        "# human_review_decision_schema_v0_7",
        "",
        f"final_state: `{FINAL_STATE}`",
        "",
        "人工审阅标签只表达候选知识是否可进入下一轮正式化候选，不产生运行时效果。",
        "",
        "## Labels",
        "",
    ]
    for item in labels:
        lines.append(f"- `{item['label']}`: runtime_effect=`{item['runtime_effect']}`")
    lines += [
        "",
        "## Required Evidence",
        "",
        "ACCEPT/CONFIRM 类标签必须填写 human_reviewer、human_reviewer_role、reviewed_at、human_review_notes、review_basis、evidence_refs、decision_confidence。",
        "任何正式化、模板化、扣分、报告口径升级动作必须进入二次审批，不在 v0.7 自动生效。",
    ]
    (ROOT / "human_review_decision_schema_v0_7.md").write_text("\n".join(lines) + "\n", encoding="utf-8")
    return schema


def write_guidance_and_plan(queue_rows, schema):
    (ROOT / "human_review_guidance_v0_7.md").write_text(
        "# human_review_guidance_v0_7\n\n"
        f"final_state: `{FINAL_STATE}`\n\n"
        "v0.7 是人工审阅工作台数据包，不接 EcoCheck runtime。\n\n"
        "## 可填写字段\n\n"
        + "\n".join(f"- `{field}`" for field in EDITABLE_FIELDS)
        + "\n\n## 不可修改字段\n\n"
        "除上述字段外，所有 source/gate/condition/scenario/open question/risk 字段均为只读证据字段，不得在工作表中直接改写。\n\n"
        "## 审阅规则\n\n"
        "- `DIVISION_CONTEXT` 只能作为召回线索，不能作为适用证据。\n"
        "- `photo_points` / `evidence_chain` 是一等字段，不能合并进备注后丢失。\n"
        "- 人工审阅不等于运行时批准；即使填写 `CONFIRM_*`，仍不得自动生成正式 permit_type、正式检查模板或自动扣分。\n"
        "- 需要改规则时填写 `NEED_RULE_FIX`，不要直接覆盖 v0.4.1 源表。\n",
        encoding="utf-8",
    )
    (ROOT / "human_review_backfill_plan_v0_7.md").write_text(
        "# human_review_backfill_plan_v0_7\n\n"
        f"final_state: `{FINAL_STATE}`\n\n"
        "本计划只设计回灌闭环，不覆盖 v0.4.1 源表。\n\n"
        "## 回灌原则\n\n"
        "- 已填写的人工审阅表作为 overlay 输入，保留 source_row_id 与 candidate_relation_id。\n"
        "- 回灌脚本应生成 review_delta_report、failure_list、review_version_manifest。\n"
        "- 原始候选表保持不可变，审阅结论以新版本 overlay 叠加。\n\n"
        "## 可进入正式化候选\n\n"
        "- `CONFIRM_APPLIES`\n"
        "- `CONFIRM_MAY_APPLY`\n"
        "- `CONFIRM_NOT_APPLY`\n\n"
        "这些标签只代表进入正式化候选池，仍需二次审批。\n\n"
        "## 仍 BLOCKS_RUNTIME\n\n"
        "- `NEED_EIA_CONFIRM`\n"
        "- `NEED_PERMIT_CONFIRM`\n"
        "- `NEED_SITE_CONFIRM`\n"
        "- `REJECT_CANDIDATE`\n"
        "- `MERGE_DUPLICATE`\n"
        "- `NEED_RULE_FIX`\n\n"
        "## 签字留痕\n\n"
        "必须保留 human_reviewer、human_reviewer_role、reviewed_at、review_basis、evidence_refs、decision_confidence。"
        "任何运行时接入、正式模板生成、自动扣分或报告口径升级均需二次审批。\n",
        encoding="utf-8",
    )


def main():
    context_rows = read_csv(ROOT / "all_context_applicability_review_v0_4_1.csv")
    permit_rows = read_csv(ROOT / "all_permit_condition_backfill_v0_4_1.csv")
    open_questions = read_csv(ROOT / "open_questions_v0_4_1.csv")
    risks = read_csv(ROOT / "risk_acceptance_queue_v0_4_1.csv")
    graph_refs = load_graph_refs()
    queue_rows = build_queue(context_rows, graph_refs)
    write_csv(ROOT / "human_review_queue_v0_7.csv", queue_rows)
    write_json(ROOT / "human_review_queue_v0_7.json", queue_rows)
    worksheet_rows = build_worksheet(queue_rows)
    schema = build_decision_schema()
    write_guidance_and_plan(queue_rows, schema)

    gate = {
        "validation_status": "PENDING_VALIDATE",
        "final_state": FINAL_STATE,
        "runtime_integration": RUNTIME_INTEGRATION,
        "input_counts": {
            "context_relations": len(context_rows),
            "permit_conditions": len(permit_rows),
            "open_questions": len(open_questions),
            "risk_queue": len(risks),
        },
        "review_queue_rows": len(queue_rows),
        "worksheet_rows": len(worksheet_rows),
        "review_priority_counts": dict(sorted({p: sum(1 for r in queue_rows if r["review_priority"] == p) for p in {r["review_priority"] for r in queue_rows}}.items())),
        "review_bucket_counts": dict(sorted({b: sum(1 for r in queue_rows if r["review_bucket"] == b) for b in {r["review_bucket"] for r in queue_rows}}.items())),
        "human_review_prefill_count": 0,
        "runtime_effect": "NONE",
    }
    write_json(ROOT / "human_review_v0_7_gate_report.json", gate)
    (ROOT / "human_review_v0_7_gate_report.md").write_text(
        "# human_review_v0_7_gate_report\n\n"
        f"- final_state: `{FINAL_STATE}`\n"
        f"- runtime_integration: `{RUNTIME_INTEGRATION}`\n"
        f"- context_relations: {len(context_rows)}\n"
        f"- permit_conditions: {len(permit_rows)}\n"
        f"- open_questions: {len(open_questions)}\n"
        f"- risk_queue: {len(risks)}\n"
        f"- review_queue_rows: {len(queue_rows)}\n"
        f"- human_review_prefill_count: 0\n",
        encoding="utf-8",
    )
    (ROOT / "FINAL_COMPLETION_REPORT_v0_7.md").write_text(
        "# FINAL COMPLETION REPORT v0.7\n\n"
        f"最终状态：`{FINAL_STATE}`\n\n"
        "v0.7 已生成人工审阅工作台数据包与审阅闭环设计，未接 EcoCheck runtime。\n\n"
        f"- human_review_queue_rows: {len(queue_rows)}\n"
        f"- worksheet_rows: {len(worksheet_rows)}\n"
        "- human_review_label / human_reviewer / reviewed_at 均未预填。\n"
        "- 不生成正式 permit_type、正式检查模板或自动扣分。\n",
        encoding="utf-8",
    )
    print(json.dumps({"final_state": FINAL_STATE, "review_queue_rows": len(queue_rows), "worksheet_rows": len(worksheet_rows)}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
